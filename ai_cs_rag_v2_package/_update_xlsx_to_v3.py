#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
更新 跨境进口家电AI客服_RAG底稿深化_v2.xlsx → v3
- 备份原 xlsx 到 _backup_20260414_154018/（如未在该目录存在）
- 更新 5 张关键 sheet：sku_master_60, real_doc_checklist, field_specs_priority5,
  rag_doc_registry, eval_batch1_priority5
- 新增 sheet：eval_batch1_v2_expanded（160 条新评测集）
- 更新 v2_summary 为 v3_summary（保留原 sheet）
- 更新 README 标注为 v3
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import csv, json, shutil, os
from datetime import datetime

WORKDIR = 'D:/个人知识库/个人知识库/AI产品经理/ai客服家电/ai_cs_rag_v2_package'
os.chdir(WORKDIR)

XLSX_PATH = '跨境进口家电AI客服_RAG底稿深化_v2.xlsx'
BACKUP_DIR = '_backup_20260414_154018'

# 备份 xlsx（若未备份）
if not os.path.exists(os.path.join(BACKUP_DIR, XLSX_PATH)):
    shutil.copy(XLSX_PATH, os.path.join(BACKUP_DIR, XLSX_PATH))
    print(f'✅ 备份 xlsx 到 {BACKUP_DIR}/')
else:
    print(f'⏭ xlsx 备份已存在，跳过')

wb = openpyxl.load_workbook(XLSX_PATH)
HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='305496')
WRAP_ALIGN = Alignment(wrap_text=True, vertical='top')

def reset_sheet(sheet_name):
    """删除并重建空 sheet，返回 ws"""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    return ws

def write_header(ws, headers):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.freeze_panes = 'A2'

# ─────────────────────────────────────
# 1. 更新 sku_master_60
# ─────────────────────────────────────
print('\n[1/6] 更新 sku_master_60')
with open('sku_master_60.csv', encoding='utf-8-sig') as f:
    reader = csv.reader(f)
    headers = next(reader)
    rows = list(reader)
ws = reset_sheet('sku_master_60')
write_header(ws, headers)
for r in rows:
    ws.append(r)
print(f'  → {len(rows)} 行 × {len(headers)} 列')

# ─────────────────────────────────────
# 2. 更新 real_doc_checklist
# ─────────────────────────────────────
print('\n[2/6] 更新 real_doc_checklist')
with open('real_doc_checklist.csv', encoding='utf-8-sig') as f:
    reader = csv.reader(f)
    headers = next(reader)
    rows = list(reader)
ws = reset_sheet('real_doc_checklist')
write_header(ws, headers)
for r in rows:
    ws.append(r)
print(f'  → {len(rows)} 行 × {len(headers)} 列')

# ─────────────────────────────────────
# 3. 更新 field_specs_priority5
# ─────────────────────────────────────
print('\n[3/6] 更新 field_specs_priority5')
with open('field_specs_priority5.json', encoding='utf-8') as f:
    field_data = json.load(f)
# field_data 可能是 list of dict 或 dict with key 'fields'
if isinstance(field_data, dict):
    fields = field_data.get('fields', field_data.get('field_specs', []))
else:
    fields = field_data

if fields:
    # 收集所有 key
    all_keys = []
    for spec in fields:
        for k in spec.keys():
            if k not in all_keys:
                all_keys.append(k)
    ws = reset_sheet('field_specs_priority5')
    write_header(ws, all_keys)
    for spec in fields:
        row = []
        for k in all_keys:
            v = spec.get(k, '')
            if isinstance(v, (list, dict)):
                v = json.dumps(v, ensure_ascii=False)
            row.append(v)
        ws.append(row)
    print(f'  → {len(fields)} 行 × {len(all_keys)} 列')
else:
    print('  ⚠ field_specs_priority5.json 解析为空')

# ─────────────────────────────────────
# 4. 更新 rag_doc_registry（从 v3 jsonl 重新生成）
# ─────────────────────────────────────
print('\n[4/6] 更新 rag_doc_registry')
docs = [json.loads(l) for l in open('rag_source_docs_v2.jsonl', encoding='utf-8')]

reg_headers = ['doc_id','doc_type','domain','title','brand','category','sku_id',
               'version_country','sales_regions','supported_use_countries','warranty_scope_type',
               'priority_topics','source_type','source_of_truth_tier','confidence_ceiling',
               'subject_risk_level','data_risk_level','locale','version','expires_at',
               'replace_with_real_doc']
ws = reset_sheet('rag_doc_registry')
write_header(ws, reg_headers)

def get_field(d, *keys, default=''):
    """支持嵌套：d['facts']['china_usability'] → get_field(d, 'facts', 'china_usability')"""
    cur = d
    for k in keys:
        if isinstance(cur, dict) and k in cur:
            cur = cur[k]
        else:
            return default
    if isinstance(cur, list):
        return '|'.join(str(x) for x in cur)
    if isinstance(cur, bool):
        return 'true' if cur else 'false'
    return str(cur) if cur is not None else default

for d in docs:
    row = [
        d.get('doc_id',''),
        d.get('doc_type',''),
        d.get('domain',''),
        d.get('title',''),
        d.get('brand') or get_field(d, 'brand_id'),
        d.get('category',''),
        d.get('sku_id',''),
        d.get('version_country',''),
        '|'.join(d.get('sales_regions',[])) if isinstance(d.get('sales_regions'),list) else d.get('sales_regions',''),
        '|'.join(d.get('supported_use_countries',[])) if isinstance(d.get('supported_use_countries'),list) else d.get('supported_use_countries',''),
        d.get('warranty_scope_type',''),
        '|'.join(d.get('priority_topics',[])) if isinstance(d.get('priority_topics'),list) else d.get('priority_topics',''),
        d.get('source_type',''),
        d.get('source_of_truth_tier',''),
        d.get('confidence_ceiling',''),
        d.get('subject_risk_level',''),
        d.get('data_risk_level',''),
        d.get('locale',''),
        d.get('version',''),
        d.get('expires_at',''),
        'true' if d.get('replace_with_real_doc') else 'false',
    ]
    ws.append(row)
print(f'  → {len(docs)} 行 × {len(reg_headers)} 列')

# ─────────────────────────────────────
# 5. 更新 eval_batch1_priority5（保留为原始视图）+ 新增 eval_batch1_v2_expanded
# ─────────────────────────────────────
print('\n[5/6] 新增 eval_batch1_v2_expanded')
with open('eval_batch1_v2_expanded.csv', encoding='utf-8-sig') as f:
    reader = csv.reader(f)
    headers = next(reader)
    rows = list(reader)
ws = reset_sheet('eval_batch1_v2_expanded')
write_header(ws, headers)
for r in rows:
    ws.append(r)
# 设置列宽并自动换行（query/rubric_3pt 可能很长）
for col_letter in ['G','U']:  # query, rubric_3pt
    ws.column_dimensions[col_letter].width = 50
print(f'  → {len(rows)} 行 × {len(headers)} 列')

# ─────────────────────────────────────
# 6. 新增 v3_summary + 更新 README
# ─────────────────────────────────────
print('\n[6/6] 新增 v3_summary + 更新 README')
ws = reset_sheet('v3_summary')
ws.append(['指标', 'v2 原始', 'v3 当前', '说明'])
for c in ws[1]:
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
v3_summary_rows = [
    ('jsonl 文档数', '220', '310', '+90 (60 import_warranty + 10 shipping_eta + 20 travel_use)'),
    ('sku_master 列数', '23', '33', '+10 (china_usability/warranty_type 等)'),
    ('field_specs 条数', '54', '78', '+24 字段规范'),
    ('real_doc_checklist 条数', '197', '287', '+90 (与 jsonl 同步)'),
    ('eval 评测样例数', '100', '160', '+60 (B1-B12 边界)'),
    ('eval 列数', '11', '23', '+12 (含 gold_must_contain 等机评字段)'),
    ('doc_type 类型数', '6', '9', '+3 (import_warranty/shipping_eta/travel_use)'),
    ('JSON Schema', '无', 'rag_schema_v3.json', 'Draft 2020-12 含 allOf 跨字段约束'),
    ('信任度隔离', '无', '4 级 source_of_truth_tier', 'tier1-4 + confidence_ceiling 0.3-1.0'),
    ('修复矛盾数', '-', '7 类', '详见 fix_notes.md'),
    ('备份位置', '-', '_backup_20260414_154018/', '原始 v2 的 5 个文件 + xlsx'),
    ('修复时间', '-', '2026-04-14 ~ 2026-04-15', '执行人：AI 产品经理'),
]
for r in v3_summary_rows:
    ws.append(r)
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 25
ws.column_dimensions['D'].width = 60
print(f'  → v3_summary 共 {len(v3_summary_rows)} 行')

# 更新 README
ws = wb['README']
# 把现有 README 内容追加 v3 说明
last_row = ws.max_row
ws.cell(row=last_row+2, column=1, value='======== v3 更新（2026-04-15）========')
ws.cell(row=last_row+2, column=1).font = Font(bold=True, color='C00000')
v3_notes = [
    'v3 修复内容详见同目录 fix_notes.md',
    'v3 评测扩充详见同目录 eval_expansion_notes.md + eval_rubric_guide.md',
    'v3 schema 详见同目录 rag_schema_v3.json (Draft 2020-12)',
    '原 v2 文件备份在 _backup_20260414_154018/ 目录',
    '本 xlsx 是镜像视图，权威源是 jsonl/csv/json 文件',
]
for i, note in enumerate(v3_notes):
    ws.cell(row=last_row+3+i, column=1, value=f'  • {note}')

# 保存
wb.save(XLSX_PATH)
print(f'\n✅ 已保存 → {XLSX_PATH}')
print(f'当前 sheet 数：{len(wb.sheetnames)}')
